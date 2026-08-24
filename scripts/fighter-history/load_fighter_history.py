#!/usr/bin/env python3
"""Load fighter-platform historical order exports into private.fighter_orders_staging.

Reads the 44-column XLSX exports from the two desktop folders (EFFEN ALL is
canonical; VISENA SDN BHD duplicates only append to source_entities), dedupes
by the platform-global Order ID, normalizes (MYT→UTC, phone digits, MY/SG,
items), and loads Supabase in batches through the Management API SQL endpoint
(auth: the CLI access token in ~/.supabase/access-token — no DB password
needed). See supabase/migrations/20260812181411_fighter_history_import.sql
for the staging DDL and the transform.

Usage:
  load               parse + load staging (idempotent: on conflict do nothing)
  load --truncate    truncate staging first (full reload)
  load --dry-run     parse + report + write CSV only, no network
  transform          run private.fighter_import_run() in chunks until done
  csv                parse + write CSV for the psql \\copy fallback

The CSV fallback (needs SUPABASE_DB_URL, session pooler):
  psql "$SUPABASE_DB_URL" -v ON_ERROR_STOP=1 \\
    -c "truncate private.fighter_orders_staging" \\
    -c "\\copy private.fighter_orders_staging (<COLUMNS>) from 'fighter_staging.csv' with (format csv, header true)"
(The exact column list is printed by the `csv` command.)
"""

import argparse
import csv
import datetime as dt
import glob
import json
import re
import sys
import time
import urllib.request
import urllib.error
from decimal import Decimal, InvalidOperation
from pathlib import Path
from zoneinfo import ZoneInfo

PROJECT_REF = "wwgtjjekhehaepbxyrij"
API_URL = f"https://api.supabase.com/v1/projects/{PROJECT_REF}/database/query"
TOKEN_PATH = Path.home() / ".supabase" / "access-token"

FOLDERS = [
    ("EFFEN ALL", "/mnt/c/Users/Nadeem/Desktop/EFFEN ALL/EFFEN ALL"),
    ("VISENA SDN BHD", "/mnt/c/Users/Nadeem/Desktop/VISENA SDN BHD"),
]

MYT = ZoneInfo("Asia/Kuala_Lumpur")

EXPECTED_HEADER = (
    "Date", "Order ID", "Invoice Number", "Status", "Channel", "Channel URL",
    "Seller ID", "Seller Name", "Seller Role", "Seller Code", "Customer Name",
    "Address", "Street Address", "Postcode", "City", "State", "Country",
    "Email", "Phone", "Product & Variations", "SKUs", "Item Count",
    "Weights (g)", "Payment Method", "Payment URL", "Payment Status",
    "Payment Paid At", "Shipping Provider", "Pushed Date", "Tracking Number",
    "Tracking Confirmation Number", "Coupon Code", "Points", "Cost",
    "Member Price", "Sales Commission", "Shipping Charge", "COD Charge",
    "Payment Gateway Charge", "Discount (Coupon)", "Discount (Dynamic)",
    "Selling Price", "Completed At", "Notes",
)

# Staging columns in load order (generated/default columns omitted).
COLUMNS = [
    ("fighter_order_id", "bigint"), ("placed_at", "timestamptz"),
    ("invoice_number", "text"), ("status", "text"), ("channel", "text"),
    ("channel_url", "text"), ("domain", "text"), ("seller_id", "text"),
    ("seller_name", "text"), ("seller_role", "text"), ("seller_code", "text"),
    ("customer_name", "text"), ("address", "text"), ("street_address", "text"),
    ("postcode", "text"), ("city", "text"), ("state", "text"),
    ("country", "text"), ("email", "text"), ("phone", "text"),
    ("products_raw", "text"), ("skus_raw", "text"), ("items", "jsonb"),
    ("item_count", "int"), ("weight_g", "numeric"), ("payment_method", "text"),
    ("payment_url", "text"), ("payment_status", "text"), ("paid_at", "timestamptz"),
    ("shipping_provider", "text"), ("pushed_date", "timestamptz"),
    ("tracking_number", "text"), ("tracking_confirmation_number", "text"),
    ("coupon_code", "text"), ("points", "numeric"), ("cost", "numeric"),
    ("member_price", "numeric"), ("sales_commission", "numeric"),
    ("shipping_charge", "numeric"), ("cod_charge", "numeric"),
    ("gateway_charge", "numeric"), ("discount_coupon", "numeric"),
    ("discount_dynamic", "numeric"), ("selling_price", "numeric"),
    ("completed_at", "timestamptz"), ("notes", "text"),
    ("source_entity", "text"), ("source_entities", "text[]"),
    ("source_file", "text"),
]

ITEM_SPLIT = re.compile(r",\s*(?=\d+x\s)")
ITEM_TOKEN = re.compile(r"^(\d+)x\s+(.*)$", re.S)


def text(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)
    s = str(v).strip()
    return s or None


def ts(v):
    """Export timestamps are MYT; emit UTC ISO-8601."""
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        d = v
    else:
        s = str(v).strip()
        if not s:
            return None
        d = dt.datetime.strptime(s, "%d/%m/%Y %H:%M:%S")
    return d.replace(tzinfo=MYT).astimezone(dt.timezone.utc).isoformat()


def num(v):
    if v is None or v == "":
        return None
    try:
        return str(Decimal(str(v)))
    except InvalidOperation:
        return None


def digits(v):
    s = text(v)
    if not s:
        return None
    d = re.sub(r"[^0-9]", "", s)
    return d or None


def host(v):
    s = text(v)
    if not s:
        return None
    s = s.lower()
    s = re.sub(r"^https?://", "", s)
    s = re.sub(r"^www\.", "", s)
    return s.strip().strip("/") or None


COUNTRY = {"malaysia": "MY", "singapore": "SG"}


def parse_items(products, skus, oid, warnings):
    """Zip 'Nx name' product tokens with 'Nx SKU' tokens positionally."""
    p_tokens = ITEM_SPLIT.split(products.strip()) if products else []
    s_tokens = ITEM_SPLIT.split(skus.strip()) if skus else []
    items = []
    if p_tokens and len(p_tokens) == len(s_tokens):
        for p, s in zip(p_tokens, s_tokens):
            pm, sm = ITEM_TOKEN.match(p.strip()), ITEM_TOKEN.match(s.strip())
            items.append({
                "sku": sm.group(2).strip() if sm else (s.strip() or None),
                "name": pm.group(2).strip() if pm else p.strip(),
                "quantity": int(sm.group(1)) if sm else (int(pm.group(1)) if pm else 1),
            })
    else:
        if p_tokens and s_tokens:
            warnings.append(f"items zip mismatch on order {oid}: {len(p_tokens)} products vs {len(s_tokens)} skus")
        for p in p_tokens:
            pm = ITEM_TOKEN.match(p.strip())
            items.append({
                "sku": None,
                "name": pm.group(2).strip() if pm else p.strip(),
                "quantity": int(pm.group(1)) if pm else 1,
            })
    return [i for i in items if i["name"]]


def parse_all():
    import openpyxl

    rows = {}          # fighter_order_id -> row dict
    warnings = []
    file_counts = []
    for entity, folder in FOLDERS:
        files = sorted(glob.glob(f"{folder}/20*/*.xlsx"))
        if not files:
            sys.exit(f"no xlsx files under {folder} — is the drive mounted?")
        for path in files:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb[wb.sheetnames[0]]
            it = ws.iter_rows(values_only=True)
            header = tuple(c if c is not None else "" for c in next(it))
            if header[: len(EXPECTED_HEADER)] != EXPECTED_HEADER:
                sys.exit(f"header drift in {path}:\n{header}")
            fname = Path(path).name
            n = 0
            for r in it:
                oid = r[1]
                if oid is None:
                    continue
                oid = int(oid)
                n += 1
                if oid in rows:
                    ents = rows[oid]["source_entities"]
                    if entity not in ents:
                        ents.append(entity)
                    continue
                placed = ts(r[0])
                if placed is None:
                    warnings.append(f"unparseable date on order {oid} in {fname}: {r[0]!r}")
                    continue
                rows[oid] = {
                    "fighter_order_id": oid,
                    "placed_at": placed,
                    "invoice_number": text(r[2]),
                    "status": text(r[3]) or "Unknown",
                    "channel": text(r[4]),
                    "channel_url": text(r[5]),
                    "domain": host(r[5]),
                    "seller_id": text(r[6]),
                    "seller_name": text(r[7]),
                    "seller_role": text(r[8]),
                    "seller_code": text(r[9]),
                    "customer_name": text(r[10]),
                    "address": text(r[11]),
                    "street_address": text(r[12]),
                    "postcode": text(r[13]),
                    "city": text(r[14]),
                    "state": text(r[15]),
                    "country": COUNTRY.get((text(r[16]) or "").lower()),
                    "email": (text(r[17]) or "").lower() or None,
                    "phone": digits(r[18]),
                    "products_raw": text(r[19]),
                    "skus_raw": text(r[20]),
                    "items": parse_items(text(r[19]), text(r[20]), oid, warnings),
                    "item_count": int(r[21]) if r[21] is not None else None,
                    "weight_g": num(r[22]),
                    "payment_method": text(r[23]),
                    "payment_url": text(r[24]),
                    "payment_status": text(r[25]),
                    "paid_at": ts(r[26]),
                    "shipping_provider": text(r[27]),
                    "pushed_date": ts(r[28]),
                    "tracking_number": text(r[29]),
                    "tracking_confirmation_number": text(r[30]),
                    "coupon_code": text(r[31]),
                    "points": num(r[32]),
                    "cost": num(r[33]),
                    "member_price": num(r[34]),
                    "sales_commission": num(r[35]),
                    "shipping_charge": num(r[36]),
                    "cod_charge": num(r[37]),
                    "gateway_charge": num(r[38]),
                    "discount_coupon": num(r[39]),
                    "discount_dynamic": num(r[40]),
                    "selling_price": num(r[41]),
                    "completed_at": ts(r[42]),
                    "notes": text(r[43]),
                    "source_entity": entity,
                    "source_entities": [entity],
                    "source_file": fname,
                }
            file_counts.append((entity, fname, n))
            wb.close()
            print(f"  {entity} | {fname} | {n} rows", flush=True)

    dual = sum(1 for r in rows.values() if len(r["source_entities"]) > 1)
    print(f"\nparsed: {len(rows)} unique orders ({dual} present in both folders)")
    for w in warnings[:20]:
        print(f"  WARN {w}")
    if len(warnings) > 20:
        print(f"  ... {len(warnings) - 20} more warnings")
    return rows


def api_query(token, sql, retries=3):
    body = json.dumps({"query": sql}).encode()
    for attempt in range(1, retries + 1):
        req = urllib.request.Request(
            API_URL, data=body, method="POST",
            headers={
                "Authorization": f"Bearer {token}",
                "Content-Type": "application/json",
                # api.supabase.com sits behind Cloudflare, which 403s (error 1010)
                # the default urllib user-agent.
                "User-Agent": "supabase-cli/2.33.9 fighter-history-loader",
            },
        )
        try:
            with urllib.request.urlopen(req, timeout=300) as resp:
                return json.loads(resp.read().decode() or "null")
        except urllib.error.HTTPError as e:
            detail = e.read().decode()[:500]
            if attempt == retries or e.code in (400, 401, 403):
                sys.exit(f"API error {e.code}: {detail}")
            print(f"  retry {attempt} after API {e.code}: {detail}", flush=True)
            time.sleep(5 * attempt)
        except (urllib.error.URLError, TimeoutError) as e:
            if attempt == retries:
                sys.exit(f"network error: {e}")
            print(f"  retry {attempt} after network error: {e}", flush=True)
            time.sleep(5 * attempt)


def dollar_quote(payload):
    for i in range(1000):
        tag = f"$fj{i}$"
        if tag not in payload:
            return tag
    sys.exit("could not find a safe dollar-quote tag")


def load(rows, token, truncate=False, batch_size=500):
    col_names = ", ".join(c for c, _ in COLUMNS)
    rec_types = ", ".join(f"{c} {t}" for c, t in COLUMNS)
    if truncate:
        print("truncating staging…")
        api_query(token, "truncate private.fighter_orders_staging")
    ordered = sorted(rows.values(), key=lambda r: r["fighter_order_id"])
    total = len(ordered)
    done = 0
    t0 = time.time()
    for i in range(0, total, batch_size):
        batch = ordered[i : i + batch_size]
        payload = json.dumps(batch, ensure_ascii=False)
        tag = dollar_quote(payload)
        sql = (
            f"insert into private.fighter_orders_staging ({col_names})\n"
            f"select {col_names} from jsonb_to_recordset({tag}{payload}{tag}::jsonb)\n"
            f"  as r({rec_types})\n"
            f"on conflict (fighter_order_id) do nothing"
        )
        api_query(token, sql)
        done += len(batch)
        if done % 10000 < batch_size or done == total:
            rate = done / max(time.time() - t0, 1)
            print(f"  loaded {done}/{total} ({rate:.0f} rows/s)", flush=True)
    check = api_query(token, "select count(*)::int as n, count(*) filter (where array_length(source_entities,1)=2)::int as dual from private.fighter_orders_staging")
    print(f"staging now: {check[0]['n']} rows, {check[0]['dual']} dual-entity")


def transform(token, in_transit="completed", chunk=10000):
    while True:
        res = api_query(
            token,
            f"set statement_timeout = '5min'; "
            f"select * from private.fighter_import_run('{in_transit}', {chunk})",
        )
        r = res[0]
        print(f"  staged={r['staged']} matched+={r['matched']} inserted+={r['inserted']} "
              f"reconciled+={r['reconciled']} pending={r['pending']}", flush=True)
        if int(r["pending"]) == 0:
            break
    print("transform complete")


def write_csv(rows, path):
    ordered = sorted(rows.values(), key=lambda r: r["fighter_order_id"])
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([c for c, _ in COLUMNS])
        for r in ordered:
            out = []
            for c, t in COLUMNS:
                v = r[c]
                if v is None:
                    out.append("")
                elif t == "jsonb":
                    out.append(json.dumps(v, ensure_ascii=False))
                elif t == "text[]":
                    out.append("{" + ",".join(f'"{e}"' for e in v) + "}")
                else:
                    out.append(v)
            w.writerow(out)
    print(f"wrote {path} ({len(ordered)} rows)")
    print("\\copy column list:\n" + ", ".join(c for c, _ in COLUMNS))


def main():
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("command", choices=["load", "transform", "csv"])
    ap.add_argument("--truncate", action="store_true", help="truncate staging before load")
    ap.add_argument("--dry-run", action="store_true", help="parse + CSV only, no network")
    ap.add_argument("--batch-size", type=int, default=500)
    ap.add_argument("--chunk", type=int, default=10000, help="transform batch limit")
    ap.add_argument("--in-transit", default="completed", choices=["completed", "processing"])
    ap.add_argument("--csv-path", default="fighter_staging.csv")
    args = ap.parse_args()

    token = None
    if not (args.command == "csv" or args.dry_run):
        if not TOKEN_PATH.exists():
            sys.exit(f"missing {TOKEN_PATH}; run `supabase login` or use the csv + psql fallback")
        token = TOKEN_PATH.read_text().strip()

    if args.command == "transform":
        transform(token, args.in_transit, args.chunk)
        return

    rows = parse_all()
    if args.command == "csv" or args.dry_run:
        write_csv(rows, args.csv_path)
        return
    load(rows, token, truncate=args.truncate, batch_size=args.batch_size)


if __name__ == "__main__":
    main()
